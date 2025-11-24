import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.styles import Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
import re
from typing import Union, List, Tuple


def create_rich_text_cell(original_text: str, substitutions: List[Tuple[str, str]]) -> Union[CellRichText, str]:
    """
    Create a rich text cell where substituted ingredients are colored red.
    
    Args:
        original_text: The full cell text
        substitutions: List of (original_ingredient, replacement_ingredient) tuples
    
    Returns:
        CellRichText object with substituted ingredients in red, or plain str if no substitutions
    """
    if not substitutions:
        return original_text
    
    # Build a list of text segments with their formatting
    segments = []
    
    # Track positions of all substitutions in the modified text
    current_text = original_text
    replacement_positions = []
    
    # Find all replacement positions in the current text
    for substitution_tuple in substitutions:
        # Handle both old format (original, replacement) and new format (original, replacement, meal_type)
        if len(substitution_tuple) == 3:
            original, replacement, _ = substitution_tuple
        else:
            original, replacement = substitution_tuple
        # Find all occurrences of the replacement text (case-insensitive)
        pattern = re.compile(re.escape(replacement), re.IGNORECASE)
        for match in pattern.finditer(current_text):
            replacement_positions.append((match.start(), match.end(), replacement))
    
    # Sort by position
    replacement_positions.sort()
    
    # Merge overlapping intervals
    merged_positions = []
    if replacement_positions:
        current_start, current_end, _ = replacement_positions[0]
        
        for next_start, next_end, _ in replacement_positions[1:]:
            if next_start < current_end:
                # Overlap, extend current end
                current_end = max(current_end, next_end)
            else:
                # No overlap, push current and start new
                merged_positions.append((current_start, current_end))
                current_start, current_end = next_start, next_end
        merged_positions.append((current_start, current_end))
    
    # Build rich text segments
    last_end = 0
    red_font = InlineFont(color='FFFF0000')  # Red color (8-digit ARGB: alpha=FF, RGB=FF0000)
    
    for start, end in merged_positions:
        # Add normal text before this replacement
        if start > last_end:
            normal_text = current_text[last_end:start]
            if normal_text:
                segments.append(normal_text)
        
        # Add red text for the replacement (only if not empty)
        replacement_text = current_text[start:end]
        if replacement_text:
            segments.append(TextBlock(red_font, replacement_text))
        last_end = end
    
    # Add any remaining normal text (only if not empty)
    if last_end < len(current_text):
        remaining_text = current_text[last_end:]
        if remaining_text:
            segments.append(remaining_text)
    
    # Create rich text object
    if segments:
        return CellRichText(*segments)
    else:
        return original_text


def export_to_excel(df: pd.DataFrame, menu_processor) -> io.BytesIO:
    """
    Export DataFrame to Excel with red highlighting for substituted ingredients.
    
    Args:
        df: The modified DataFrame to export
        menu_processor: MenuProcessor instance with substitution_map
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    
    # Write data to worksheet with rich text formatting
    for row_idx in range(len(df)):
        for col_idx in range(len(df.columns)):
            cell_value = df.iloc[row_idx, col_idx]
            
            # Excel cells are 1-indexed
            excel_row = row_idx + 1
            excel_col = col_idx + 1
            
            if pd.notna(cell_value):
                # Get substitutions for this cell
                substitutions = menu_processor.get_substitutions_for_cell(row_idx, col_idx)
                
                if substitutions:
                    # Create rich text with red highlighting
                    rich_text = create_rich_text_cell(str(cell_value), substitutions)
                    ws.cell(row=excel_row, column=excel_col).value = rich_text
                else:
                    # Plain text
                    ws.cell(row=excel_row, column=excel_col).value = str(cell_value)
                
                # Set alignment to wrap text and align top-left
                ws.cell(row=excel_row, column=excel_col).alignment = Alignment(
                    wrap_text=True,
                    vertical='top',
                    horizontal='left'
                )
    
    # Auto-adjust column widths (set to reasonable default)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 50
    
    # Save to BytesIO buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
